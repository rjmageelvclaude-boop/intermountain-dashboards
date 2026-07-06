#!/usr/bin/env python3
"""
Hyperion dashboard data pipeline (v2).

Reads the two ServiceTitan / Enterprise Hub exports and writes site/hyperion-data.js.

    py build/parse_reports.py [sales.xlsx] [installed.xlsx]

Defaults to the copies in data/. Re-run whenever new reports arrive; in production a
scheduled job pulls the emailed reports from Gmail and runs this automatically.
"""
import sys, os, json, datetime
from collections import defaultdict
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
CONFIG = json.load(open(os.path.join(HERE, "config.json"), encoding="utf-8"))
OUT = os.path.join(ROOT, "site", "hyperion-data.js")

SALES = sys.argv[1] if len(sys.argv) > 1 else os.path.join(ROOT, "data", "sales_by_rep.xlsx")
INST  = sys.argv[2] if len(sys.argv) > 2 else os.path.join(ROOT, "data", "installed_by_rep.xlsx")

TEAMS = CONFIG["teams"]
NON_CONTEST = set(CONFIG.get("nonContestTenants", []))
WEEKS = CONFIG["contest"]["weeks"]

# tenant -> [team keys] (crisafulli maps to two, disambiguated by business unit)
TENANT_TEAMS = defaultdict(list)
for key, t in TEAMS.items():
    for tenant in t["tenants"]:
        TENANT_TEAMS[tenant].append(key)


def team_for(tenant, business_unit):
    keys = TENANT_TEAMS.get(tenant)
    if not keys:
        return None
    if len(keys) == 1:
        return keys[0]
    bu = business_unit or ""
    for k in keys:
        f = TEAMS[k].get("buFilter")
        if f and f.lower() in bu.lower():
            return k
    return keys[0]  # unmatched BU: default to first (flagged team split)


def read_sheet(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    hdr = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        d = dict(zip(hdr, row))
        if not d.get("Tenant Name"):  # report's embedded grand-total row
            continue
        yield d


def read_date_range(path):
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        for r in wb["Filters"].iter_rows(values_only=True):
            if r and r[0] == "Date Range":
                return r[1]
    except Exception:
        pass
    return ""


def num(v):
    return float(v) if isinstance(v, (int, float)) else 0.0


def day(v):
    return v.date() if isinstance(v, datetime.datetime) else v if isinstance(v, datetime.date) else None


def first_name(techs):
    return (techs or "").split(",")[0].strip()


# ---------------- load rows ----------------
inst_rows, sales_rows, exceptions = [], [], []
max_date = datetime.date(2026, 1, 1)

for d in read_sheet(INST):
    tenant = d["Tenant Name"]
    team = team_for(tenant, d.get("Business Unit"))
    rep = (d.get("Sold By") or "").strip()
    dt = day(d.get("Completion Date")) or day(d.get("Scheduled Date"))
    amt = num(d.get("Jobs Total"))
    if dt and dt > max_date:
        max_date = dt
    row = {"team": team, "tenant": tenant, "rep": rep, "date": dt, "amt": amt}
    inst_rows.append(row)
    if not rep and amt != 0:
        exceptions.append({
            "team": TEAMS[team]["name"] if team else tenant,
            "tenant": tenant,
            "ours": bool(team and TEAMS[team].get("ours")),
            "job": str(d.get("Job #") or ""),
            "invoice": str(d.get("Invoice #") or ""),
            "customer": (d.get("Customer Name") or "").strip(),
            "bu": d.get("Business Unit") or "",
            "jobType": d.get("Job Type") or "",
            "amount": round(amt),
            "completed": dt.isoformat() if dt else "",
        })

for d in read_sheet(SALES):
    tenant = d["Tenant Name"]
    team = team_for(tenant, d.get("Business Unit"))
    rep = first_name(d.get("Assigned Technicians"))
    dt = day(d.get("Scheduled Date")) or day(d.get("Created Date"))
    amt = num(d.get("Jobs Estimate Sales Subtotal"))
    if dt and dt > max_date:
        max_date = dt
    sales_rows.append({"team": team, "tenant": tenant, "rep": rep, "date": dt, "amt": amt})

as_of = max_date
start = datetime.date.fromisoformat(CONFIG["contest"]["startDate"])
day_num = max(1, (as_of - start).days + 1)

cur_week = WEEKS[-1]
for w in WEEKS:
    if datetime.date.fromisoformat(w["start"]) <= as_of <= datetime.date.fromisoformat(w["end"]):
        cur_week = w
        break
wk_start = datetime.date.fromisoformat(cur_week["start"])
wk_end = datetime.date.fromisoformat(cur_week["end"])
days_left_week = max(0, (wk_end - as_of).days)


def bucket(rows, keyfn):
    """-> {key: {today,wtd,contest}} using each row's date."""
    out = defaultdict(lambda: {"today": 0.0, "wtd": 0.0, "contest": 0.0})
    for r in rows:
        k = keyfn(r)
        if k is None:
            continue
        b = out[k]
        b["contest"] += r["amt"]
        if r["date"]:
            if r["date"] >= wk_start:
                b["wtd"] += r["amt"]
            if r["date"] == as_of:
                b["today"] += r["amt"]
    return out


# ---------------- teams ----------------
inst_team = bucket(inst_rows, lambda r: r["team"])
sales_team = bucket(sales_rows, lambda r: r["team"])
unattr_team = bucket([r for r in inst_rows if not r["rep"]], lambda r: r["team"])

# roster = distinct reps seen in either report (overridable)
roster = defaultdict(set)
for r in inst_rows + sales_rows:
    if r["team"] and r["rep"]:
        roster[r["team"]].add(r["rep"])
overrides = CONFIG.get("rosterOverrides", {})

teams_out = {}
for key, t in TEAMS.items():
    n_ca = overrides.get(key) or len(roster[key]) or 1
    inst = inst_team.get(key, {"today": 0, "wtd": 0, "contest": 0})
    sales = sales_team.get(key, {"today": 0, "wtd": 0, "contest": 0})
    teams_out[key] = {
        "key": key, "name": t["name"], "division": t["division"],
        "ours": bool(t.get("ours")), "color": t.get("color", "#6B7A8C"),
        "caCount": n_ca,
        "installed": {k: round(v) for k, v in inst.items()},
        "sales": {k: round(v) for k, v in sales.items()},
        "revPerCaWeek": round(inst["wtd"] / n_ca),
        "revPerCaContest": round(inst["contest"] / n_ca),
        "unattributed": round(unattr_team.get(key, {}).get("contest", 0)),
        "wins": 0, "losses": 0, "points": 0, "gmBonus": False,
    }

# ---------------- reps (merged installed + sales) ----------------
inst_rep = bucket([r for r in inst_rows if r["rep"]], lambda r: (r["team"] or r["tenant"], r["rep"]))
sales_rep = bucket([r for r in sales_rows if r["rep"]], lambda r: (r["team"] or r["tenant"], r["rep"]))

reps_out = []
for (tk, rep) in set(inst_rep) | set(sales_rep):
    team = teams_out.get(tk)
    i = inst_rep.get((tk, rep), {"today": 0, "wtd": 0, "contest": 0})
    s = sales_rep.get((tk, rep), {"today": 0, "wtd": 0, "contest": 0})
    reps_out.append({
        "name": rep,
        "teamKey": tk if team else None,
        "company": team["name"] if team else tk,
        "ours": team["ours"] if team else False,
        "color": team["color"] if team else "#6B7A8C",
        "installed": {k: round(v) for k, v in i.items()},
        "sales": {k: round(v) for k, v in s.items()},
    })
reps_out.sort(key=lambda r: -r["installed"]["contest"])

exceptions.sort(key=lambda e: (-e["amount"]))

data = {
    "generatedAt": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
    "dateRange": read_date_range(INST) or read_date_range(SALES),
    "asOf": as_of.isoformat(),
    "dayNum": day_num,
    "daysLeftWeek": days_left_week,
    "week": {"num": cur_week["num"], "label": cur_week["label"]},
    "phase": "Regular Season" if cur_week["num"] <= 5 else "Postseason",
    "contest": {k: v for k, v in CONFIG["contest"].items() if k != "weeks"},
    "matchups": CONFIG["matchups"].get(str(cur_week["num"]), []),
    "teams": teams_out,
    "reps": reps_out,
    "exceptions": exceptions,
}

os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, "w", encoding="utf-8") as f:
    f.write("// AUTO-GENERATED by build/parse_reports.py — do not edit by hand.\n")
    f.write("window.HYPERION_DATA = ")
    json.dump(data, f, indent=1)
    f.write(";\n")

n_ours = sum(1 for t in teams_out.values() if t["ours"])
print(f"Wrote {OUT}")
print(f"  asOf {as_of} (day {day_num}/62, {cur_week['label']}), {len(teams_out)} teams ({n_ours} ours), {len(reps_out)} reps")
print(f"  installed ${sum(t['installed']['contest'] for t in teams_out.values()):,} | sales ${sum(t['sales']['contest'] for t in teams_out.values()):,}")
print(f"  exceptions: {len(exceptions)} jobs, ${sum(e['amount'] for e in exceptions):,} unattributed")
